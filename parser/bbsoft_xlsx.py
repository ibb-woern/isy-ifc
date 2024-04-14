import os
from typing import List, Union
import enum
from typing import Type

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from models.types import (
    ManholeFormType,
    ProfileType,
    StatusType,
    SystemType,
    MaterialType,
)
from models.manhole import Manhole
from models.pipe_section import PipeSection


def _try_match_enum(value: str, enum_type: Type[enum.Enum]) -> Union[enum.Enum, None]:
    try:
        return enum_type(value)
    except ValueError:
        return None


def _create_row_dict(row, headers) -> dict:
    row_dict = {}
    i = 0
    for column in row:
        row_dict[headers[i]] = column
        i += 1
    return row_dict


def _find_manhole_by_name(name: str, manholes: List[Manhole]) -> Union[Manhole, None]:
    pass


def _manholes(ws: Worksheet) -> List[Manhole]:
    manholes = []
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for row in ws.iter_rows(values_only=True, min_row=2):
        row_dict = _create_row_dict(row, headers)
        manhole = Manhole(
            name=row_dict["ID_CUR"],
            x=row_dict["POS_X"],
            y=row_dict["POS_Y"],
            z=row_dict["POS_Z"],
            z_top=row_dict["POS_ZLOW"],
            depth=row_dict["POS_DZ"],
            status=_try_match_enum(row_dict["STATUS.KZ"], StatusType),
            system=_try_match_enum(row_dict["TYP.KZ"], SystemType),
            form=_try_match_enum(row_dict["FORM.KZ"], ManholeFormType),
            coverplate=row_dict["COVERPLATE.KZ"],
            cone=row_dict["CONE.KZ"],
            nominal_length=row_dict["SIZE_A"],
            nominal_width=row_dict["SIZE_B"],
            material=_try_match_enum(row_dict["MATERIAL.KZ"], MaterialType),
        )
        manholes.append(manhole)
    return manholes


def _get_millimeter_value(data: dict, key: str) -> Union[float, None]:
    if key in data.keys():
        if data[key] is None:
            return None
        return data[key] / 1000
    return None


def _pipe_sections(ws: Worksheet, manholes: List[Manhole]) -> List[PipeSection]:
    sewers = []
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for row in ws.iter_rows(values_only=True, min_row=2):
        row_dict = _create_row_dict(row, headers)

        diameter_inner = _get_millimeter_value(row_dict, "HYD_PROFH")
        wall_thickness = _get_millimeter_value(row_dict, "WALL_THICKNESS")
        diameter_outer = None
        if diameter_inner is not None and wall_thickness is not None:
            diameter_outer = diameter_inner + 2 * wall_thickness

        sewer = PipeSection(
            name=row_dict["ID_NAME"],
            start=_find_manhole_by_name(row_dict["ID_DRAIN1"], manholes),
            end=_find_manhole_by_name(row_dict["ID_DRAIN2"], manholes),
            x_1=row_dict["ABS_1X"],
            y_1=row_dict["ABS_1Y"],
            z_1=row_dict["ABS_1Z"],
            x_2=row_dict["ABS_2X"],
            y_2=row_dict["ABS_2Y"],
            z_2=row_dict["ABS_2Z"],
            status=_try_match_enum(row_dict["STATUS.KZ"], StatusType),
            system=_try_match_enum(row_dict["HYD_PRINCIP.KZ"], SystemType),
            profile=_try_match_enum(row_dict["HYD_PROF.KZ"], ProfileType),
            diameter_inner=diameter_inner,
            diameter_outer=diameter_outer,
            material=_try_match_enum(row_dict["OTH_MATERIAL.KZ"], MaterialType),
        )
        sewers.append(sewer)

    return sewers


def parse(file_path: os.PathLike):
    wb = load_workbook(file_path, data_only=True)
    manholes: List[Manhole] = None
    sewers: List[PipeSection] = None

    # iterate over sheets
    for ws in wb.worksheets:
        if ws.title == "schaechte":
            manholes = _manholes(ws)
        if ws.title == "haltungen":
            sewers = _pipe_sections(ws, manholes)
        if ws.title == "leitungen":
            print("Leitungen are currently not yet supported")
        if ws.title == "strasseneinlaeufe":
            print("Strasseneinlaeufe are currently not yet supported")
        if ws.title == "hausanschluesse":
            print("Hausanschluesse are currently not yet supported")
        if ws.title == "bauwerke":
            print("Bauwerke are currently not yet supported")

    return manholes, sewers
